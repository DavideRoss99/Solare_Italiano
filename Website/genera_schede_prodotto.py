"""
Genera le schede HTML da incollare nelle descrizioni prodotto di Squarespace Commerce.
Output: una cartella Website/schede_html/ con un file HTML per ogni prodotto.

Uso: python3 genera_schede_prodotto.py
"""

import openpyxl
import os
import re
from pathlib import Path

XLSX_PATH = Path(__file__).parent / '0. Scheda Prodotti.xlsx'
OUTPUT_DIR = Path(__file__).parent / 'schede_html'

# Mappa combustibile → CSS tag class + emoji + etichetta
FUEL_MAP = {
    'legna':     ('si-fuel-tag--legna',     '🌲', 'Legna'),
    'pellet':    ('si-fuel-tag--pellet',    '🔥', 'Pellet'),
    'combinato': ('si-fuel-tag--combinato', '🔥🌲', 'Pellet + Legna'),
    'gas':       ('si-fuel-tag--gas',       '🔵', 'Gas'),
    'caldaia':   ('si-fuel-tag--caldaia',   '🌡️', 'Caldaia'),
    'pdc':       ('si-fuel-tag--pdc',       '♻️', 'Pompa di Calore'),
    'solare':    ('si-fuel-tag--solare',    '☀️', 'Solare'),
    'ev':        ('si-fuel-tag--ev',        '⚡', 'Ricarica EV'),
}

def detect_fuel_class(combustione: str) -> tuple:
    if not combustione:
        return ('si-fuel-tag--gas', '•', combustione or '')
    c = str(combustione).lower()
    if 'pellet' in c and 'legna' in c:
        return FUEL_MAP['combinato']
    if 'pellet' in c:
        return FUEL_MAP['pellet']
    if 'legna' in c:
        return FUEL_MAP['legna']
    if 'gas' in c:
        return FUEL_MAP['gas']
    if 'pompa' in c or 'calore' in c or 'pdc' in c:
        return FUEL_MAP['pdc']
    if 'solare' in c or 'fotovoltaico' in c or 'pannelli' in c:
        return FUEL_MAP['solare']
    if 'ev' in c or 'ricarica' in c or 'colonnina' in c:
        return FUEL_MAP['ev']
    return ('si-fuel-tag--gas', '•', str(combustione))

def v(val) -> str:
    """Normalizza valore cella: None → 'n.d.' """
    if val is None:
        return 'n.d.'
    s = str(val).strip()
    return s if s else 'n.d.'

def safe_filename(name: str) -> str:
    return re.sub(r'[^a-zA-Z0-9_\-]', '_', str(name).strip())[:80]

def spec_row(label: str, value, highlight=False) -> str:
    if value is None or str(value).strip() == '':
        val_str = 'n.d.'
    else:
        val_str = str(value).strip()
    cls = ' si-specs__row--highlight' if highlight else ''
    return f'''      <tr class="si-specs__row{cls}">
        <td class="si-specs__label">{label}</td>
        <td class="si-specs__value">{val_str}</td>
      </tr>'''

def section_header(title: str) -> str:
    return f'      <tr><td colspan="2" class="si-specs__section-header">{title}</td></tr>'

def make_html(row: tuple, header: tuple) -> str:
    # Map header to dict
    d = {str(h).replace('\n', ' ').strip(): v(c) for h, c in zip(header, row)}

    cat        = d.get('Categoria Prodotto', '')
    combustione= d.get('Combustione', '')
    marchio    = d.get('Marchio', '')
    nome       = d.get('Nome Prodotto', '')
    desc_corta = d.get('Descrizione Corta', '')
    desc_lunga = d.get('Descrizione Lunga', '')
    prezzo     = d.get('Prezzo (€)', '')
    potenza    = d.get('Potenza Nominale (kW)', '')
    mq         = d.get('Riscaldamento max (m²)', '')
    serbatoio  = d.get('Serbatoio Pellet (kg)', '')
    dimensioni = d.get('Dimensioni B×H×P (mm)', '')
    rendimento = d.get('Rendimento (%)', '')
    classe_en  = d.get('Classe Energetica', '')
    cert       = d.get('Certificazioni', '')
    ppm        = d.get('PPM (mg/Nm³)', '')
    canal      = d.get('Canalizzazione', '')
    pulizia    = d.get('Sistema Pulizia', '')
    wifi       = d.get('Wi-Fi', '')
    colori     = d.get('Colori', '')
    varianti   = d.get('Altre Varianti', '')

    fuel_class, fuel_emoji, fuel_label = detect_fuel_class(combustione)

    # Badges
    badge_power = f'''    <div class="si-badge si-badge--power">
      <span class="si-badge__icon">🔥</span>
      <span class="si-badge__label">Potenza</span>
      <span class="si-badge__value">{potenza} kW</span>
    </div>''' if potenza != 'n.d.' else ''

    badge_area = f'''    <div class="si-badge si-badge--area">
      <span class="si-badge__icon">📐</span>
      <span class="si-badge__label">Riscalda fino a</span>
      <span class="si-badge__value">{mq} m²</span>
    </div>''' if mq != 'n.d.' else ''

    badge_fuel = f'''    <div class="si-badge si-badge--fuel">
      <span class="si-badge__icon">{fuel_emoji}</span>
      <span class="si-badge__label">Combustibile</span>
      <span class="si-badge__value">{fuel_label}</span>
    </div>'''

    badge_price = f'''    <div class="si-badge si-badge--price">
      <span class="si-badge__icon">💶</span>
      <span class="si-badge__label">Prezzo (IVA esclusa)</span>
      <span class="si-badge__value">€ {prezzo}</span>
    </div>''' if prezzo not in ('n.d.', '') else ''

    badges_html = '\n'.join(b for b in [badge_power, badge_area, badge_fuel, badge_price] if b)

    # Descrizione corta (prosa)
    desc_html = ''
    if desc_corta and desc_corta != 'n.d.':
        desc_html = f'  <p class="si-description">{desc_corta}</p>\n'

    # Spec table rows
    rows = [
        section_header('Prestazioni'),
        spec_row('Potenza nominale (kW)', potenza),
        spec_row('Rendimento', f'{rendimento}%' if rendimento != 'n.d.' else 'n.d.'),
        spec_row('Riscaldamento max (m²)', mq),
        spec_row('Classe energetica', classe_en),
    ]

    if serbatoio not in ('n.d.', 'N/A', ''):
        rows += [
            section_header('Pellet'),
            spec_row('Serbatoio pellet', f'{serbatoio} kg'),
        ]

    rows += [
        section_header('Caratteristiche tecniche'),
        spec_row('Dimensioni B×H×P (mm)', dimensioni),
        spec_row('Canalizzazione', canal),
        spec_row('Sistema pulizia', pulizia),
        spec_row('Wi-Fi', wifi),
    ]

    if ppm not in ('n.d.', ''):
        rows.append(spec_row('Emissioni PPM (mg/Nm³)', ppm))

    rows += [
        section_header('Finiture e varianti'),
        spec_row('Colori', colori),
        spec_row('Varianti disponibili', varianti),
        section_header('Certificazioni'),
        spec_row('Certificazioni', cert, highlight=True),
    ]

    rows_html = '\n'.join(rows)

    html = f'''<!-- SOLARE ITALIANO — Scheda prodotto: {marchio} {nome} -->
<!-- Incolla questo HTML nel campo Descrizione del prodotto su Squarespace Commerce -->
<!-- CSS necessario: Design → Custom CSS (file squarespace_custom.css) -->

<span class="si-fuel-tag {fuel_class}">{fuel_emoji} {fuel_label}</span>

<div class="si-spec-badges">
{badges_html}
</div>

{desc_html}<details class="si-specs">
  <summary class="si-specs__trigger">
    <span class="si-specs__trigger-label">
      <svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M9 11l3 3L22 4"/><path d="M21 12v7a2 2 0 01-2 2H5a2 2 0 01-2-2V5a2 2 0 012-2h11"/></svg>
      Specifiche tecniche
    </span>
    <span class="si-specs__arrow">
      <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24"><polyline points="6 9 12 15 18 9"/></svg>
    </span>
  </summary>
  <div class="si-specs__body">
    <table class="si-specs__table">
{rows_html}
    </table>
  </div>
</details>

<a href="/contattaci" class="si-cta">
  Richiedi informazioni o sopralluogo gratuito →
</a>'''

    return html


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        header = rows[2]  # riga 3 = intestazioni colonne
        data_rows = [r for r in rows[3:] if r[3] is not None]  # col 4 = Nome Prodotto

        if not data_rows:
            continue

        brand_dir = OUTPUT_DIR / safe_filename(sheet_name)
        brand_dir.mkdir(exist_ok=True)

        for row in data_rows:
            marchio = str(row[2]).strip() if row[2] else sheet_name
            nome    = str(row[3]).strip() if row[3] else 'senza_nome'
            fname   = f"{safe_filename(marchio)}_{safe_filename(nome)}.html"

            html = make_html(row, header)
            (brand_dir / fname).write_text(html, encoding='utf-8')
            print(f"  ✓ {sheet_name}/{fname}")
            total += 1

    print(f"\nGenerati {total} file HTML in {OUTPUT_DIR}/")
    print("\nCome usarli:")
    print("  1. Apri Squarespace → Negozio → seleziona prodotto")
    print("  2. Nel campo Descrizione, clicca '<>' (HTML source)")
    print("  3. Incolla il contenuto del file HTML corrispondente")
    print("  4. Salva")


if __name__ == '__main__':
    main()
